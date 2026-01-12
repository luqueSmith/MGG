import re
import hashlib
from pathlib import Path
from io import BytesIO

import requests
from PIL import Image as PILImage

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ========= CONFIG =========
DATA_FILE = Path("js/data-mutants.js")      # <- aquí está window.RAW_LIST
OUT_XLSX  = Path("mutantes_con_imagenes.xlsx")

BASE = "https://s-ak.kobojo.com/mutants/assets/thumbnails/"
TIMEOUT = 6
RETRIES = 2
IMG_SIZE = 64

CACHE_DIR = Path(".img_cache_mutants")
CACHE_DIR.mkdir(exist_ok=True)


# ========= PARSER =========
def extract_raw_list(js_text: str) -> str:
    """
    Extrae el contenido dentro de: window.RAW_LIST = `...`;
    """
    m = re.search(r"window\.RAW_LIST\s*=\s*`([\s\S]*?)`\s*;", js_text)
    if not m:
        raise RuntimeError("No encontré window.RAW_LIST dentro de js/data-mutants.js")
    return m.group(1)

def parse_rows(raw: str):
    """
    Cada línea es: Nombre<TAB>codigo
    """
    rows = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        name = parts[0].strip()
        code = parts[1].strip().lower()
        rows.append((name, code))
    return rows


# ========= IMAGES =========
def url_platinum(code: str) -> str:
    return f"{BASE}specimen_{code}_platinum.png"

def url_normal(code: str) -> str:
    return f"{BASE}specimen_{code}.png"

def cache_path_for_url(url: str) -> Path:
    h = hashlib.sha1(url.encode("utf-8")).hexdigest()
    return CACHE_DIR / f"{h}.png"

def download_bytes(url: str) -> bytes | None:
    """
    Descarga con timeout + retries. Usa cache.
    """
    cp = cache_path_for_url(url)
    if cp.exists():
        return cp.read_bytes()

    headers = {"User-Agent": "Mozilla/5.0"}
    last_err = None

    for _ in range(RETRIES + 1):
        try:
            r = requests.get(url, timeout=TIMEOUT, headers=headers)
            if r.status_code == 200 and r.content:
                cp.write_bytes(r.content)
                return r.content
            last_err = f"HTTP {r.status_code}"
        except Exception as e:
            last_err = str(e)

    return None

def make_square_thumb(img_bytes: bytes, size: int) -> BytesIO:
    im = PILImage.open(BytesIO(img_bytes)).convert("RGBA")
    w, h = im.size
    s = min(w, h)
    left = (w - s) // 2
    top  = (h - s) // 2
    im = im.crop((left, top, left + s, top + s)).resize((size, size), PILImage.LANCZOS)
    out = BytesIO()
    im.save(out, format="PNG")
    out.seek(0)
    return out

def best_image_bytes(code: str):
    """
    Prueba platinum y si falla, normal.
    """
    b = download_bytes(url_platinum(code))
    if b:
        return b, url_platinum(code), url_normal(code)
    b = download_bytes(url_normal(code))
    return b, url_platinum(code), url_normal(code)


# ========= EXCEL =========
def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def build_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mutantes"

    headers = ["Imagen", "Nombre", "Código", "Thumb Platinum", "Thumb Normal"]
    ws.append(headers)

    # estilos header
    header_fill = PatternFill("solid", fgColor="1B4332")  # verde oscuro
    header_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # anchos
    set_col_width(ws, 1, 12)
    set_col_width(ws, 2, 34)
    set_col_width(ws, 3, 16)
    set_col_width(ws, 4, 62)
    set_col_width(ws, 5, 62)

    # filas
    missing = 0
    total = len(rows)
    for idx, (name, code) in enumerate(rows, start=2):
        if (idx - 1) % 25 == 0:
            print(f"…progreso: {idx-1}/{total}")

        img_bytes, u_plat, u_norm = best_image_bytes(code)

        ws.cell(row=idx, column=2, value=name)
        ws.cell(row=idx, column=3, value=code)
        ws.cell(row=idx, column=4, value=u_plat)
        ws.cell(row=idx, column=5, value=u_norm)

        # formato alineación
        for c in range(2, 6):
            ws.cell(row=idx, column=c).alignment = Alignment(vertical="center")
        ws.row_dimensions[idx].height = 54

        # zebra suave
        if idx % 2 == 0:
            fill = PatternFill("solid", fgColor="0B0506")
            for c in range(1, 6):
                ws.cell(row=idx, column=c).fill = fill
                ws.cell(row=idx, column=c).font = Font(color="FFFaf0")

        # insertar imagen
        if img_bytes:
            thumb = make_square_thumb(img_bytes, IMG_SIZE)
            ximg = XLImage(thumb)
            ximg.anchor = f"A{idx}"
            ws.add_image(ximg)
        else:
            missing += 1
            ws.cell(row=idx, column=1, value="SIN IMG").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=idx, column=1).font = Font(bold=True, color="C1121F")

    print(f"✅ Terminado. Sin imagen: {missing}/{total}")
    return wb


def main():
    if not DATA_FILE.exists():
        raise RuntimeError(f"No existe {DATA_FILE}. Ajusta la ruta DATA_FILE.")

    js_text = DATA_FILE.read_text(encoding="utf-8")
    raw = extract_raw_list(js_text)
    rows = parse_rows(raw)

    print(f"✅ Mutantes encontrados: {len(rows)}")
    wb = build_excel(rows)
    wb.save(OUT_XLSX)
    print(f"🎉 Excel creado: {OUT_XLSX.resolve()}")

if __name__ == "__main__":
    main()
