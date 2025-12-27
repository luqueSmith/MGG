import re
from pathlib import Path
from io import BytesIO

import requests
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DATA_FILE = Path("js/mutants.js")
OUT_XLSX = Path("mutantes_con_imagenes.xlsx")

BASE = "https://s-ak.kobojo.com/mutants/assets/thumbnails/"
TIMEOUT = 5       # ⬅️ clave: no esperar infinito
IMG_SIZE = 64

def parse_mutants(text):
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("//"):
            continue
        parts = re.split(r"\t+", line)
        if len(parts) >= 2:
            rows.append((parts[0].strip(), parts[1].strip().lower()))
    return rows

def fetch_image(url):
    try:
        r = requests.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        return r.content
    except Exception:
        return None

def make_thumb(img_bytes):
    im = PILImage.open(BytesIO(img_bytes)).convert("RGBA")
    w, h = im.size
    s = min(w, h)
    im = im.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2))
    im = im.resize((IMG_SIZE, IMG_SIZE))
    out = BytesIO()
    im.save(out, format="PNG")
    out.seek(0)
    return out

print("📄 Leyendo mutantes...")
text = DATA_FILE.read_text(encoding="utf-8")
mutants = parse_mutants(text)
print(f"✅ {len(mutants)} mutantes encontrados\n")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mutantes"

ws.append(["Imagen", "Nombre", "Código"])
ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 16

for i, (name, code) in enumerate(mutants, start=2):
    print(f"⬇️  [{i-1}/{len(mutants)}] {name} ({code})")

    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=code)
    ws.row_dimensions[i].height = 52

    url_platinum = f"{BASE}specimen_{code}_platinum.png"
    url_normal = f"{BASE}specimen_{code}.png"

    img = fetch_image(url_platinum) or fetch_image(url_normal)
    if img:
        thumb = make_thumb(img)
        xl_img = XLImage(thumb)
        xl_img.anchor = f"A{i}"
        ws.add_image(xl_img)

wb.save(OUT_XLSX)
print(f"\n🎉 LISTO: {OUT_XLSX.resolve()}")
